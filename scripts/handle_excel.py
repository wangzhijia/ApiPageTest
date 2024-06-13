from collections import namedtuple

from openpyxl import load_workbook

from configs.constants import CASES_FILE_PATH, CONFIG_FILE_PATH
from scripts.handle_config import HandleConfig


class HandleExcel:

    def __init__(self, file_name, sheet_name=None):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.file_name)
        self.ws = self.wb[self.sheet_name]
        if self.sheet_name is None:
            self.ws = self.wb.active
        else:
            self.ws = self.wb[self.sheet_name]
        self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        self.Cases = namedtuple('Cases', self.sheet_head_tuple)
        self.case_list = []

    def get_cases(self):
        for data in self.ws.iter_rows(min_row=2, values_only=True):
            self.case_list.append(self.Cases(*data))
        return self.case_list

    def write_cases(self, row, actual, result):
        config = HandleConfig(CONFIG_FILE_PATH)
        other_wb = load_workbook(self.file_name)
        other_ws = other_wb[self.sheet_name]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=config.get_int('excel', 'actual_col'), value=actual)
            other_ws.cell(row=row, column=config.get_int('excel', 'result_col'), value=result)
            other_wb.save(self.file_name)
        else:
            print("传入的行号有误，行号应为大于1的整数")


if __name__ == '__main__':
    handle_excel = HandleExcel(CASES_FILE_PATH, 'toutiao_index')
    cases_list = handle_excel.get_cases()
    print(cases_list)
    print(len(cases_list))













